"""
search_documents.py - Search Excel documentation files for table/column info.

Reads .xlsx/.xls files from the documents/ folder, searches for keywords across
sheet names, column headers, and cell values.  Caches parsed data as JSON for
fast repeat searches.

Usage:
    python search_documents.py --keyword "khach hang"
    python search_documents.py --keyword "doanh thu" --folder documents/
    python search_documents.py --keyword "REVENUE" --no-cache
    python search_documents.py --keyword "CONTRACT" --format json
"""

import argparse
import hashlib
import json
import os
import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


# ============================================================================
# Cache
# ============================================================================

CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".doc_cache")


def _file_hash(filepath: str) -> str:
    """Return MD5 of file content for cache invalidation."""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _cache_path(filepath: str) -> str:
    name = Path(filepath).stem
    return os.path.join(CACHE_DIR, f"{name}.json")


def _load_cache(filepath: str) -> list[dict] | None:
    """Load cached parsed data if still valid."""
    cp = _cache_path(filepath)
    if not os.path.exists(cp):
        return None
    try:
        with open(cp, "r", encoding="utf-8") as f:
            cached = json.load(f)
        if cached.get("file_hash") == _file_hash(filepath):
            return cached["rows"]
    except (json.JSONDecodeError, KeyError):
        pass
    return None


def _save_cache(filepath: str, rows: list[dict]):
    os.makedirs(CACHE_DIR, exist_ok=True)
    cp = _cache_path(filepath)
    with open(cp, "w", encoding="utf-8") as f:
        json.dump({"file_hash": _file_hash(filepath), "rows": rows}, f, ensure_ascii=False)


# ============================================================================
# Parsers
# ============================================================================

def _parse_xlsx(filepath: str) -> list[dict]:
    """Parse .xlsx file into list of {file, sheet, row_num, headers, values}."""
    if openpyxl is None:
        print("openpyxl is required for .xlsx files. Install: pip install openpyxl", file=sys.stderr)
        return []

    rows = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            str_values = [str(v) if v is not None else "" for v in row]
            if row_idx == 1:
                headers = str_values
                continue
            rows.append({
                "file": os.path.basename(filepath),
                "sheet": sheet_name,
                "row_num": row_idx,
                "headers": headers,
                "values": str_values,
            })
    wb.close()
    return rows


def _parse_xls(filepath: str) -> list[dict]:
    """Parse .xls file into list of dicts."""
    if xlrd is None:
        print("xlrd is required for .xls files. Install: pip install xlrd", file=sys.stderr)
        return []

    rows = []
    wb = xlrd.open_workbook(filepath)
    for sheet in wb.sheets():
        headers = []
        for row_idx in range(sheet.nrows):
            str_values = [str(sheet.cell_value(row_idx, c)) for c in range(sheet.ncols)]
            if row_idx == 0:
                headers = str_values
                continue
            rows.append({
                "file": os.path.basename(filepath),
                "sheet": sheet.name,
                "row_num": row_idx + 1,
                "headers": headers,
                "values": str_values,
            })
    return rows


def parse_file(filepath: str, use_cache: bool = True) -> list[dict]:
    """Parse an Excel file (with optional caching)."""
    if use_cache:
        cached = _load_cache(filepath)
        if cached is not None:
            return cached

    ext = Path(filepath).suffix.lower()
    if ext == ".xlsx":
        rows = _parse_xlsx(filepath)
    elif ext == ".xls":
        rows = _parse_xls(filepath)
    else:
        print(f"Unsupported file format: {ext}", file=sys.stderr)
        return []

    if use_cache:
        _save_cache(filepath, rows)

    return rows


# ============================================================================
# Search
# ============================================================================

def search_documents(keyword: str, folder: str = "documents/",
                     use_cache: bool = True, use_regex: bool = False,
                     limit: int = 200) -> list[dict]:
    """
    Search all Excel files in folder for keyword.

    Returns list of matches with context: file, sheet, row_num, matched_field, context.
    """
    # Resolve folder relative to skill root
    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    abs_folder = os.path.join(skill_root, folder) if not os.path.isabs(folder) else folder

    if not os.path.isdir(abs_folder):
        print(f"Thư mục không tồn tại: {abs_folder}", file=sys.stderr)
        return []

    # Collect all Excel files
    excel_files = []
    for f in os.listdir(abs_folder):
        if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$"):
            excel_files.append(os.path.join(abs_folder, f))

    if not excel_files:
        print(f"Không tìm thấy file Excel trong: {abs_folder}", file=sys.stderr)
        return []

    results = []
    for fp in sorted(excel_files):
        rows = parse_file(fp, use_cache=use_cache)

        # Search in sheet names (reported once per sheet)
        seen_sheets = set()
        for row in rows:
            sheet = row["sheet"]
            if sheet not in seen_sheets and _match(keyword, sheet, use_regex):
                seen_sheets.add(sheet)
                results.append({
                    "file": row["file"], "sheet": sheet,
                    "row_num": None, "matched_field": "sheet_name",
                    "context": f"Sheet name: {sheet}",
                })

        # Search in headers (reported once per sheet)
        seen_headers = set()
        for row in rows:
            sheet = row["sheet"]
            for hi, h in enumerate(row["headers"]):
                key = (sheet, hi)
                if key not in seen_headers and _match(keyword, h, use_regex):
                    seen_headers.add(key)
                    results.append({
                        "file": row["file"], "sheet": sheet,
                        "row_num": 1, "matched_field": "header",
                        "context": f"Header [{hi}]: {h}",
                    })

        # Search in cell values
        for row in rows:
            for vi, v in enumerate(row["values"]):
                if _match(keyword, v, use_regex):
                    header = row["headers"][vi] if vi < len(row["headers"]) else f"col_{vi}"
                    # Build context: show the whole row as key=value pairs
                    context_parts = []
                    for i, val in enumerate(row["values"]):
                        if val:
                            hdr = row["headers"][i] if i < len(row["headers"]) else f"col_{i}"
                            context_parts.append(f"{hdr}={val}")
                    context_str = " | ".join(context_parts[:8])  # limit to 8 fields

                    results.append({
                        "file": row["file"], "sheet": row["sheet"],
                        "row_num": row["row_num"],
                        "matched_field": f"cell:{header}",
                        "context": context_str,
                    })

            if len(results) >= limit:
                break
        if len(results) >= limit:
            break

    return results[:limit]


def _match(pattern: str, text: str, use_regex: bool) -> bool:
    if not text:
        return False
    if use_regex:
        return bool(re.search(pattern, text, re.IGNORECASE))
    return pattern.lower() in text.lower()


# ============================================================================
# Formatters
# ============================================================================

def format_text(results: list[dict]) -> str:
    if not results:
        return "Không tìm thấy kết quả nào."

    lines = [f"Tìm thấy {len(results)} kết quả:\n"]
    lines.append(f"{'FILE':<30} {'SHEET':<25} {'ROW':<6} {'MATCH TYPE':<15} {'CONTEXT'}")
    lines.append("-" * 140)
    for r in results:
        lines.append(
            f"{r['file']:<30} {r['sheet']:<25} {str(r['row_num'] or ''):<6} "
            f"{r['matched_field']:<15} {r['context'][:90]}"
        )
    return "\n".join(lines)


def format_json(results: list[dict]) -> str:
    return json.dumps(results, indent=2, ensure_ascii=False)


def format_markdown(results: list[dict]) -> str:
    if not results:
        return "Không tìm thấy kết quả nào."
    lines = [f"Tìm thấy **{len(results)}** kết quả:\n"]
    lines.append("| File | Sheet | Row | Match | Context |")
    lines.append("|------|-------|-----|-------|---------|")
    for r in results:
        ctx = (r["context"] or "")[:80].replace("|", "\\|")
        lines.append(
            f"| {r['file']} | {r['sheet']} | {r['row_num'] or ''} "
            f"| {r['matched_field']} | {ctx} |"
        )
    return "\n".join(lines)


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Search Excel documentation files for table/column information."
    )
    parser.add_argument("--keyword", "-k", required=True, help="Search keyword or regex")
    parser.add_argument("--folder", default="documents/", help="Folder containing Excel files")
    parser.add_argument("--no-cache", action="store_true", help="Disable caching")
    parser.add_argument("--regex", action="store_true", help="Use regex matching")
    parser.add_argument("--limit", type=int, default=200, help="Max results (default: 200)")
    parser.add_argument("--format", choices=["text", "json", "markdown"], default="text")
    args = parser.parse_args()

    try:
        results = search_documents(
            keyword=args.keyword, folder=args.folder,
            use_cache=not args.no_cache, use_regex=args.regex, limit=args.limit,
        )
        if args.format == "json":
            print(format_json(results))
        elif args.format == "markdown":
            print(format_markdown(results))
        else:
            print(format_text(results))
    except Exception as e:
        print(f"Lỗi: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
